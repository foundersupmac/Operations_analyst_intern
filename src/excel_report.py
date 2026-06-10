"""Automated weekly Excel report (week 17): 6 formatted sheets with an
embedded chart, generated with openpyxl. Designed to be scheduled
(Task Scheduler / cron)."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

import config
from src.etl_pipeline import read_table
from src.kpi_framework import oee_components

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def write_sheet(wb, title, df):
    ws = wb.create_sheet(title)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for col in ws.columns:
        width = max(len(str(c.value)) for c in col if c.value is not None)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 30)
    return ws


def main():
    production = oee_components(read_table("production"))
    downtime = read_table("downtime")
    last_week = production["date"].max() - pd.Timedelta(days=6)
    week = production[production["date"] >= last_week]

    wb = Workbook()
    wb.remove(wb.active)

    summary = pd.DataFrame({
        "Metric": ["Units produced", "Defects", "Avg OEE %", "Avg availability %"],
        "Value": [int(week["units_produced"].sum()), int(week["defect_count"].sum()),
                  round(100 * week["oee"].mean(), 1),
                  round(100 * week["availability"].mean(), 1)],
    })
    write_sheet(wb, "Summary", summary)

    by_line = (week.groupby("line")
               .agg(units=("units_produced", "sum"), defects=("defect_count", "sum"),
                    oee_pct=("oee", lambda s: round(100 * s.mean(), 1)))
               .reset_index())
    ws = write_sheet(wb, "By Line", by_line)
    chart = BarChart()
    chart.title = "OEE % by line (last 7 days)"
    chart.add_data(Reference(ws, min_col=4, min_row=1, max_row=len(by_line) + 1),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=len(by_line) + 1))
    ws.add_chart(chart, "F2")

    write_sheet(wb, "By Shift", (week.groupby("shift")
                .agg(units=("units_produced", "sum"),
                     defect_rate_pct=("defect_count", "sum")).reset_index()))
    dt_week = downtime[downtime["date"] >= last_week]
    write_sheet(wb, "Downtime", dt_week.groupby("cause_code")["duration_hrs"]
                .agg(["count", "sum"]).round(1).reset_index())
    write_sheet(wb, "Top Defect Days", week.nlargest(10, "defect_count")
                [["date", "line", "shift", "defect_count"]])
    write_sheet(wb, "Data Quality", pd.DataFrame({
        "Check": ["Rows this week", "Null defect counts", "Lines reporting"],
        "Value": [len(week), int(week["defect_count"].isna().sum()),
                  week["line"].nunique()],
    }))

    path = config.REPORTS_DIR / "weekly_operations_report.xlsx"
    wb.save(path)
    print(f"  weekly Excel report written: {path}")
    return path


if __name__ == "__main__":
    main()
